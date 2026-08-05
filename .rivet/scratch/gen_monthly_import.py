# -*- coding: utf-8 -*-
"""从"总维修明细"按月拆分生成维修记录导入 Excel（临时探针脚本，收尾清理）。

列映射（总维修明细 col -> OriginalRecord entity 27 列）：
  日期(2)->日期, 班次(3)->班次, 厂房(4)->厂房, 序列号(5)->序号, 机台号(6)->机台号,
  诊断人(7)->诊断人, 维修人(8)->维修人, 报修时间(9)->报修时间, 开始时间(10)->开始时间,
  结束时间(11)->结束时间, 维修工时(12), 停机工时(13), 机型(14), 故障现象(15),
  处理方式(16)->维修描述, 料号(17)->物料编码, 配件名称(18)->零件名称, 数量(19),
  上机物料(20), 下机物料(21), 备注(22), 送货记录(23)->送货记录引用。
  确认人/单据号/上次上机时间/是否过保：源表无此列，留空（由后端计算）。
  年+月(0) 原样写入（后端按 recordDate 重算覆盖）。
"""
import datetime
import os
from collections import OrderedDict

import openpyxl

EXCEL = '金属厂--2607--截止2026年8月1日最新版.xlsx'
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'monthly_import')

HEADERS = ["年+月", "日期", "班次", "厂房", "序号", "机台号", "诊断人", "维修人",
           "报修时间", "开始时间", "结束时间", "维修工时", "停机工时", "机型", "故障现象",
           "维修描述", "物料编码", "零件名称", "数量", "上机物料", "下机物料", "备注",
           "确认人", "送货记录引用", "单据号", "上次上机时间", "是否过保"]
# 源列 -> 目标列 index（目标列按 HEADERS 顺序）
MAP = {0: 0, 2: 1, 3: 2, 4: 3, 5: 4, 6: 5, 7: 6, 8: 7, 9: 8, 10: 9, 11: 10,
       12: 11, 13: 12, 14: 13, 15: 14, 16: 15, 17: 16, 18: 17, 19: 18,
       20: 19, 21: 20, 22: 21, 23: 23}


def xl_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)):
        return datetime.date(1899, 12, 31) + datetime.timedelta(days=int(v) - 1)
    s = str(v).strip()
    try:
        return datetime.date.fromisoformat(s[:10])
    except ValueError:
        return None


def norm_time(v, date):
    """把时间值归一为完整 datetime（补日期部分）。返回 datetime 或 None。"""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, datetime.time):
        return datetime.datetime.combine(date, v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace('/', '-')
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S",
                "%H:%M:%S", "%H:%M"):
        try:
            dt = datetime.datetime.strptime(s, fmt)
            if dt.year <= 1900:
                dt = datetime.datetime.combine(date, dt.time())
            return dt
        except ValueError:
            continue
    return None


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb['总维修明细']
    rows = ws.iter_rows(values_only=True)
    next(rows)  # 表头

    groups = OrderedDict()
    skipped = 0
    for r in rows:
        d = xl_date(r[2])
        if d is None:
            skipped += 1
            continue
        key = d.strftime('%Y-%m')
        groups.setdefault(key, []).append((d, r))

    print(f"总行 {sum(len(v) for v in groups.values()) + skipped}，跳过无日期 {skipped}，"
          f"月份 {list(groups.keys())}")

    for key, items in groups.items():
        out = os.path.join(OUT_DIR, f'维修记录_{key}.xlsx')
        ow = openpyxl.Workbook()
        ows = ow.active
        ows.title = '维修记录'
        ows.append(HEADERS)
        for d, r in items:
            row = [None] * len(HEADERS)
            for src, dst in MAP.items():
                v = r[src]
                if src in (9, 10, 11):  # 报修/开始/结束时间：补全日期
                    v = norm_time(v, d)
                elif src == 2:
                    v = d
                row[dst] = v
            ows.append(row)
        ow.save(out)
        print(f"{key}: {len(items)} 行 -> {out}")


if __name__ == '__main__':
    main()
