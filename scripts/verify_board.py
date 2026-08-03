# -*- coding: utf-8 -*-
"""
数据看板对账脚本：调用后端 5 个看板接口，与 Excel 原表看板缓存值逐格对比。

用法：
  python verify_board.py                # 全量对账（company 13, year 2026）
  python verify_board.py --company 1    # 指定公司
  python verify_board.py --year 2026    # 指定年份

对比范围：
  - 维修金额/故障频次：Excel 4313 行 × 12 月 + 小计；接口机台清单 = original_record 去重（未补导，
    允许 Excel 有接口无的机台缺口，单独报告）
  - 物料频次/返修频次：183 行 × 12 月 + 合计 + 金额
  - 返修率：183 行 × 12 月 + 平均（浮点容差 1e-4）
Excel 缓存值为空的行跳过（不判错）。
"""
import argparse
import hmac
import hashlib
import base64
import json
import re
import sys
import time

import openpyxl
import requests

sys.stdout.reconfigure(encoding='utf-8')

EXCEL = r'金属厂--2607--截止2026年8月1日最新版.xlsx'
BASE = 'http://localhost:8080/api/board'
TOL = 1e-4


def make_token():
    cfg = open('backend/src/main/resources/application.yml', encoding='utf-8').read()
    secret = re.search(r'jwt:\s*\n\s+secret:\s*(\S+)', cfg).group(1)

    def b64url(d):
        return base64.urlsafe_b64encode(d).rstrip(b'=').decode()

    now = int(time.time())
    h = b64url(json.dumps({'alg': 'HS256'}).encode())
    p = b64url(json.dumps({'userId': 2, 'username': '18720647482', 'realName': 'admin',
                           'role': 'admin', 'sub': '18720647482', 'iat': now, 'exp': now + 3600}).encode())
    s = b64url(hmac.new(secret.encode(), f'{h}.{p}'.encode(), hashlib.sha256).digest())
    return f'{h}.{p}.{s}'


def fetch_board(name, company_id, year):
    r = requests.get(f'{BASE}/{name}', params={'companyId': company_id, 'year': year},
                     headers={'Authorization': f'Bearer {make_token()}'}, timeout=60)
    d = r.json()
    if d.get('code') != 200:
        raise RuntimeError(f'{name}: {r.text[:200]}')
    return d['data']


def load_excel_sheet(name):
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    return rows


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def compare_cell(api_val, xl_val, tol=TOL):
    if xl_val is None or xl_val == '':
        return None  # Excel 空值跳过
    a = 0.0 if api_val is None else float(api_val)
    b = float(xl_val)
    return abs(a - b) > tol


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--company', type=int, default=13)
    ap.add_argument('--year', type=int, default=2026)
    args = ap.parse_args()
    cid, year = args.company, args.year

    # ---- 1. 机台看板：维修金额 / 故障频次 ----
    for name, api_name in [('维修金额', 'repair-amount'), ('故障频次', 'fault-frequency')]:
        xl = load_excel_sheet(name)
        header = xl[1]
        months = [str(h) for h in header[3:15]]
        rows = fetch_board(api_name, cid, year)
        api_map = {r['key']: r for r in rows}
        # Excel 行：列 1=厂房+机台号, 3..14=12月, 15=小计
        mism = 0
        checked = 0
        missing_in_api = 0
        for line in xl[2:]:
            if line[1] is None:
                continue
            key = str(line[1]).strip()
            if key not in api_map:
                missing_in_api += 1
                continue
            row = api_map[key]
            for i, m in enumerate(months):
                res = compare_cell(row['months'].get(m), line[3 + i])
                if res is True:
                    mism += 1
                    checked += 1
                elif res is None:
                    pass
                else:
                    checked += 1
            res = compare_cell(row.get('total'), line[15])
            if res is True:
                mism += 1
                checked += 1
            elif res is None:
                pass
            else:
                checked += 1
        print(f'[{name}] 接口行 {len(rows)} | Excel 行 {len(xl)-2} | 比对 {checked} 格 | 不一致 {mism} | '
              f'Excel 有机台接口无 {missing_in_api}（未补导允许）')

    # ---- 2. 料号看板：物料频次 / 返修频次 ----
    for name, api_name in [('物料频次', 'material-frequency'), ('返修频次', 'repair-frequency')]:
        xl = load_excel_sheet(name)
        header = xl[1]
        months = [str(h) for h in header[5:17]]
        rows = fetch_board(api_name, cid, year)
        api_map = {r['key']: r for r in rows}
        mism = 0
        checked = 0
        missing = 0
        for line in xl[2:]:
            if line[2] is None:
                continue
            key = str(line[2]).strip()
            if key not in api_map:
                missing += 1
                continue
            row = api_map[key]
            for i, m in enumerate(months):
                res = compare_cell(row['months'].get(m), line[5 + i])
                if res is True:
                    mism += 1
                    checked += 1
                elif res is None:
                    pass
                else:
                    checked += 1
            # 合计列(17) 与 金额列(18)
            for ci, field in [(17, 'total'), (18, 'amount')]:
                res = compare_cell(row.get(field), line[ci])
                if res is True:
                    mism += 1
                    checked += 1
                elif res is None:
                    pass
                else:
                    checked += 1
        print(f'[{name}] 接口行 {len(rows)} | 比对 {checked} 格 | 不一致 {mism} | 接口无此料号 {missing}')

    # ---- 3. 返修率 ----
    name = '物料返修率-156'
    xl = load_excel_sheet(name)
    header = xl[1]
    months = [str(h) for h in header[5:17]]
    rows = fetch_board('repair-rate', cid, year)
    api_map = {r['key']: r for r in rows}
    mism = 0
    checked = 0
    for line in xl[2:]:
        if line[2] is None:
            continue
        key = str(line[2]).strip()
        row = api_map.get(key)
        if row is None:
            continue
        for i, m in enumerate(months):
            res = compare_cell(row['months'].get(m), line[5 + i])
            if res is True:
                mism += 1
                checked += 1
            elif res is None:
                pass
            else:
                checked += 1
        res = compare_cell(row.get('average'), line[17])
        if res is True:
            mism += 1
            checked += 1
        elif res is None:
            pass
        else:
            checked += 1
    print(f'[{name}] 比对 {checked} 格 | 不一致 {mism}')

    # ---- 4. 返修频次差异分析：T0 非空行的判别特征 ----
    print('\n=== 返修频次判别分析（T0 非空行特征）===')
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb['3.未过保物料']
    header = [str(h).strip() if h else '' for h in next(ws.iter_rows(values_only=True))]
    t0_col = header.index('辅助列-T0')
    repair_on_col = header.index('维修物料装上')
    debug_col = header.index('设备维修调试')
    rows = ws.iter_rows(values_only=True)
    next(rows)
    n_t0 = n_t0_repair_on = n_t0_debug = n_t0_both = 0
    n_all_repair_on = 0
    for row in rows:
        t0 = row[t0_col]
        repair_on = row[repair_on_col]
        debug = row[debug_col]
        if t0 is not None and str(t0).strip():
            n_t0 += 1
            if repair_on is not None and str(repair_on).strip():
                n_t0_repair_on += 1
            if debug is not None and str(debug).strip():
                n_t0_debug += 1
            if (repair_on is not None and str(repair_on).strip()) and (debug is not None and str(debug).strip()):
                n_t0_both += 1
        if repair_on is not None and str(repair_on).strip():
            n_all_repair_on += 1
    print(f'T0 非空 {n_t0} 行：其中维修物料装上非空 {n_t0_repair_on}、设备维修调试非空 {n_t0_debug}、两者都非空 {n_t0_both}')
    print(f'全表维修物料装上非空 {n_all_repair_on} 行')


if __name__ == '__main__':
    main()
